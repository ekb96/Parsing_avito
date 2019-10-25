# Данный скрипт осуществляет парсинг Avito.
# Получает заголовок объявления, цену, дату/время, ссылку.
# Запись результата в файл Excel/SQL;

# Необходимо добавить:
# проверку с заданными промежутками времени;
# добавление при появлении новых объявлений;
# сортировку результатов по времени добавления;
# оповещение на Email о новых объявлениях.

import requests
from bs4 import BeautifulSoup															# модуль для работы с html
import re																				# регулярные выражения
import openpyxl																			# библиотека для Excel
import datetime																			# модуль для работы с датой

def get_html(url):
	r = requests.get(url)
	return r.text
	
def get_total_pages(html):
	now = datetime.datetime.now()														# получаем дату
	#print ("Текущий год: %d" % now.year)												# отделяем год
	#print ("Текущий месяц: %d"% now.month)												# отделяем месяц
	#print ("Текущий день: %d" % now.day)												# отделяем день
	
	index_pages = 0																		# начальный индекс
	soup = BeautifulSoup(html, 'lxml')
																						# получаем список
	pages = soup.find_all ('div', attrs = {'class': 'description item_table-description'})
	
	pages_len = len(pages)																# кол-во элементов в списке
	#print(str(pages_len) + '\n')
	
	try:
		wb = openpyxl.load_workbook('base.xlsx')										# читаем excel-файл
	except IOError:																		# если файла нет
		wb = openpyxl.Workbook()														# создаем новый excel-файл
		wb.create_sheet(title = 'Объявления', index = 0)
	
	sheet = wb['Объявления']															# получаем лист, с которым будем работать
	
	while (index_pages != pages_len):													# цикл обхода списка
		pages1 = pages[index_pages]														# получаем элемент списка

		title = pages1.find ('span', attrs = {'itemprop': 'name'})						# достаем тег "name"
		price = pages1.find ('span', attrs = {'class': 'price'})						# достаем тег "price"
		date = pages1.find ('div', class_ = 'js-item-date c-2').get('data-absolute-date')
		href = pages1.find('a', class_ = 'item-description-title-link').get('href')		# достаем ссылку
	
		title = str(title.get_text())													# достаем заголовок
		price = str(price.get_text())													# достаем цену
		date = str(date)																# достаем дату и время
		href = "http://avito.ru" + str(href)											# дописываем ссылку
		
		title = re.sub("^\s+|\n|\r|\s+$", '', title)									# удаляем лишние символы
		price = re.sub("^\s+|\n|\r|₽|\s+$", '', price)									# удаляем лишние символы
		date = re.sub("^\s+|\n|\r|\s+$", '', date)										# удаляем лишние символы
		
		if date.find("Сегодня") == 0:
			date1 = date.split(' ')														# разделяем строку по ' '
			date1[0] = "%d" % now.day + "." + "%d" % now.month + "." + "%d" % now.year
			date = date1[0] + " " + date1[1]
			print(date)
		elif date.find("Вчера") == 0:
			date1 = date.split(' ')														# разделяем строку по ' '
			date1[0] = str(int("%d" % now.day)-1) + "." + "%d" % now.month + "." + "%d" % now.year
			date = date1[0] + " " + date1[1]
			print(date)
		else:
			print(date)
		
		print(title)
		print(price)
		print(href + '\n')
		
		sheet.append([date, title, price, href])										# запись в файл
		index_pages += 1																# индекс следующего элемента
	
	wb.save('base.xlsx')																# сохраняем файл
	
def main():
	url = "https://www.avito.ru/ekaterinburg/avtomobili/s_probegom/inomarki/mehanika/benzin/levyy_rul/ne_bolee_dvuh/ne_bityy?cd=1&pmax=350000&pmin=150000&radius=200&s=104&user=1&f=188_901b20303"
	
	html = get_html(url)
	get_total_pages(html)

if __name__ == '__main__':
	main()
